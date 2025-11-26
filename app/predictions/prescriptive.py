from pathlib import Path
import pandas as pd
from datetime import datetime
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
from app.utils.oracle_db import execute_query
from app.configs.prescriptive_conf import prescriptive_config
from app.utils.oracle_db import fetch_all
from app.configs.base_conf import settings
from datetime import datetime, timedelta
from osisoft.pidevclub.piwebapi.pi_web_api_client import PIWebApiClient
from osisoft.pidevclub.piwebapi.models import PIAnalysis, PIItemsStreamValues, PIStreamValues, PITimedValue, PIRequest
warnings.filterwarnings('ignore')

def run_prescriptive(task):
    config = prescriptive_config(task['PARAMS'])
    excel_path = config['FMEA_PATH']
    threshold_file = config['THRESHOLD_PATH']
    unit = "U" + task['PARAMS']

    output_dir = Path(config["OUTPUT_DIR"]) if config["OUTPUT_DIR"] else Path.cwd()
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Dictionary untuk menyimpan variabel dan parameternya
    components = {}

    df_fmea = load_excel_data(unit, excel_path)
    
    variable_list = load_variables_from_file(df_fmea, output_dir)

    print(variable_list)
    
    load_threshold_data(unit, threshold_file, variable_list, components)

    print("\n" + "="*80)
    print("EXTRACTED VARIABLES:")
    print("="*80)
    for idx, var in enumerate(variable_list, 1):
        print(f"{idx}. {var}")
    print(f"\nTotal: {len(variable_list)} variables")

    items = prepare_data(config, task)
    item_filterd = []
    for item in items:
        if item["NAME"] in config["VARIABLE_LIST"]:
            item_filterd.append(item)

    input = {item["NAME"]: [item["MAX_VALUE"]] for item in item_filterd}
    # print(input)

    result = process_input(input, components, df_fmea)

    upload_trip = result["summary"]["trip_count"]
    upload_alarm = result["summary"]["alarm_count"]

    upload_failure_mode_1 = ""
    upload_item_identification_1 = ""
    upload_level_1 = ""
    upload_procedure_1 = ""
    upload_recommendation_1 = ""

    upload_failure_mode_2 = ""
    upload_item_identification_2 = ""
    upload_level_2 = ""
    upload_procedure_2 = ""
    upload_recommendation_2 = ""

    upload_failure_mode_3 = ""
    upload_item_identification_3 = ""
    upload_level_3 = ""
    upload_procedure_3 = ""
    upload_recommendation_3 = ""

    save_to_excel(result, config["OUTPUT_DIR"], filename=f"monitoring_report_{task['ID']}.xlsx")
    print("TRIGGEREDS: ", len(result["triggered_fmea"]))

    # save_to_excel(result, config["OUTPUT_DIR"])
    print("Prediction completed successfully.")

    return "OKE"

def load_variables_from_file(df_fmea, output_dir):
        """
        Load variabel dari file txt (extracted_variables.txt)
        Jika file tidak ada, ekstrak dari FMEA dan buat file baru
        
        Returns:
        --------
        list : List variabel dari file (normalized)
        """
        variables_file = output_dir / "extracted_variables.txt"
        
        # Cek apakah file sudah ada
        if variables_file.exists():
            try:
                with open(variables_file, "r", encoding="utf-8") as f:
                    variables = [line.strip() for line in f.readlines() if line.strip()]
                print(f"✅ Loaded {len(variables)} variables from: {variables_file}")
                return variables
            except Exception as e:
                print(f"❌ Error reading variables file: {str(e)}")
                print("Extracting variables from FMEA instead...")
        
        # Jika file tidak ada atau error, ekstrak dari FMEA
        print("📋 File not found. Extracting variables from FMEA...")
        return extract_variables_from_fmea(df_fmea, output_dir)

def extract_variables_from_fmea(df_fmea, output_dir):
        """
        Ekstrak semua variabel unik dari kolom 'Related Variable(s)' di file FMEA
        Simpan ke file txt dan return list variabel
        
        Returns:
        --------
        list : List variabel unik yang diekstrak (normalized)
        """
        all_variables = set()
        
        if df_fmea is None:
            print("❌ Warning: FMEA data not loaded. Skipping variable extraction.")
            return []
        
        try:
            # Loop melalui kolom Related Variable(s)
            for row_idx in range(len(df_fmea)):
                related_vars = df_fmea.iloc[row_idx]["Related Variable(s)"]
                
                # Parse variabel dengan pemisah multiple
                vars_list = parse_related_variables(related_vars)
                
                # Tambahkan ke set (otomatis unique)
                for var in vars_list:
                    if var and var.lower() != "none":
                        all_variables.add(var)
        
        except Exception as e:
            print(f"❌ Error extracting variables: {str(e)}")
            return []
        
        # Sort alphabetically
        sorted_variables = sorted(all_variables)
        
        # Simpan ke file txt
        try:
            output_file = output_dir / "extracted_variables.txt"
            with open(output_file, "w", encoding="utf-8") as f:
                for var in sorted_variables:
                    f.write(f"{var}\n")
            
            print(f"✅ Extracted {len(sorted_variables)} variables from FMEA")
            print(f"📄 Saved to: {output_file}")
        
        except Exception as e:
            print(f"❌ Error saving variables to file: {str(e)}")
        
        return sorted_variables

def parse_related_variables(cell_str):
    """
    Parse Related Variable(s) dari cell string
    Pemisah bisa: newline (\n), semicolon (;), atau comma (,)
    
    Parameters:
    -----------
    cell_str : str
        String dari cell Related Variable(s)
    
    Returns:
    --------
    list : List variabel yang sudah dinormalisasi
    """
    if pd.isna(cell_str):
        return []
    
    cell_str = str(cell_str)
    
    # Split by multiple separators: newline, semicolon, comma
    # Ganti semicolon dan comma dengan newline dulu
    cell_str = cell_str.replace(";", "\n").replace(",", "\n")
    
    # Split by newline
    lines = [line.strip() for line in cell_str.split('\n') if line.strip()]
    
    # Normalize setiap variabel
    normalized_vars = [normalize_variable_name(var) for var in lines]
    
    return normalized_vars

def normalize_variable_name(var_name):
    """
    Normalize nama variabel:
    - Convert ke lowercase
    - Replace space dengan underscore
    - Strip whitespace di awal/akhir
    
    Parameters:
    -----------
    var_name : str
        Nama variabel original
    
    Returns:
    --------
    str : Nama variabel yang sudah dinormalisasi
    """
    if not isinstance(var_name, str):
        return var_name
    
    # Strip whitespace di awal dan akhir
    var_name = var_name.strip()
    
    # Convert ke lowercase
    var_name = var_name.lower()
    
    # Replace space dengan underscore
    var_name = var_name.replace(" ", "_")
    
    return var_name
        
    
def load_threshold_data(unit, threshold_file, variable_list, components):
    """Load dan proses data threshold dari file Excel sesuai unit"""
    try:
        # Tentukan sheet name berdasarkan unit
        sheet_name = f"Threshold {unit}"
        
        print(f"Loading threshold from sheet: {sheet_name}")
        
        # Baca file threshold dengan sheet yang sesuai
        df_threshold = pd.read_excel(threshold_file, sheet_name=sheet_name)
        
        # Default values untuk variabel yang tidak ada di threshold
        default_config = {
            "unit": "mm/s",
            "alarm": 65,
            "trip": 80
        }
        
        # Populate components dari variable_list
        for var_name in variable_list:
            # Cari variabel di threshold file (case-insensitive)
            threshold_row = df_threshold[
                df_threshold["TAG NAME"].str.lower().str.replace(" ", "_") == var_name
            ]
            
            if len(threshold_row) > 0:
                # Variabel ada di threshold file
                row = threshold_row.iloc[0]
                alarm = float(row["ALARM"])
                trip = float(row["TRIP"])
                
                components[var_name] = {
                    "unit": row["SATUAN"] if pd.notna(row["SATUAN"]) else default_config["unit"],
                    "alarm": alarm,
                    "trip": trip
                }
            else:
                # Variabel tidak ada di threshold file, gunakan default
                print(f"⚠️  Variabel '{var_name}' tidak ditemukan di threshold. Menggunakan default.")
                components[var_name] = default_config.copy()
                
    except FileNotFoundError:
        print(f"❌ Sheet '{sheet_name}' tidak ditemukan di file threshold")
        # Set default values jika ada error
        for var_name in variable_list:
            components[var_name] = {
                "unit": "mm/s",
                "alarm": 65,
                "trip": 80
            }
    except Exception as e:
        print(f"❌ Error loading threshold data: {str(e)}")
        # Set default values jika ada error
        for var_name in variable_list:
            components[var_name] = {
                "unit": "mm/s",
                "alarm": 65,
                "trip": 80
            }
    
def load_excel_data(unit, excel_path):
    """Load data dari Excel FMEA sesuai dengan unit yang dipilih"""
    try:
        # Tentukan sheet name berdasarkan unit
        sheet_name = f"FMEA {unit}"
        
        print(f"Loading FMEA from sheet: {sheet_name}")
        
        # Baca file Excel dengan ketentuan: header di baris 8, skip baris 9, data mulai baris 10
        df = pd.read_excel(
            excel_path,
            sheet_name=sheet_name,
            header=7,  # 0-based index, jadi 7 untuk baris 8
            skiprows=[8]  # Skip baris 9 (0-based index 8)
        )
        df_fmea = df.ffill()  # Mengatasi sel NaN karena merged
        return df_fmea
    except FileNotFoundError:
        print(f"❌ Error: File tidak ditemukan: {excel_path}")
        df_fmea = None
    except Exception as e:
        print(f"❌ Error membaca Excel: {str(e)}")
        df_fmea = None

def process_input(input_data, components, df_fmea):
        """
        Process input dataframe dan generate output
        Input dapat berupa dictionary dengan nama variabel yang belum dinormalisasi
        
        Parameters:
        -----------
        input_data : dict
            Dictionary dengan format {"var_name": [list angka], ...}
            var_name akan dinormalisasi secara otomatis
        
        Returns:
        --------
        dict : Dictionary dengan hasil analisis
            {
                "status_variables": DataFrame,
                "triggered_fmea": DataFrame,
                "overall_status": str,
                "summary": dict
            }
        """
        # Validasi input
        if not isinstance(input_data, dict):
            raise ValueError("Input harus berupa dictionary")
        
        # Normalize input_data keys
        normalized_input = {}
        for var_name, values in input_data.items():
            normalized_var = normalize_variable_name(var_name)
            normalized_input[normalized_var] = values
        
        # Extract max values dari setiap variabel
        max_values = {}
        for var_name, values in normalized_input.items():
            if var_name in components:
                if isinstance(values, list) and len(values) > 0:
                    max_values[var_name] = max(values)
                else:
                    max_values[var_name] = values
        
        # Evaluasi status setiap variabel
        variable_status = []
        trip_vars = []
        alarm_vars = []
  
        for var_name, max_val in max_values.items():
            if var_name in components:
                params = components[var_name]
                status = evaluate_crisp(max_val, params)
                
                variable_status.append({
                    "Variable": var_name,
                    "Max_Value": max_val,
                    "Unit": params["unit"],
                    "Alarm_Threshold": params["alarm"],
                    "Trip_Threshold": params["trip"],
                    "Status": status
                })
                
                if status == "Trip":
                    trip_vars.append(var_name)
                elif status == "Alarm":
                    alarm_vars.append(var_name)
        
        df_status = pd.DataFrame(variable_status)
        
        # Evaluasi FMEA triggers dengan logika AND
        triggered_fmea = []
        
        if df_fmea is not None:
            for row_idx in range(9, 247):  # Baris 10-247 (index 9-246)
                if row_idx >= len(df_fmea):
                    break
                
                related_vars = df_fmea.iloc[row_idx]["Related Variable(s)"]
                
                if pd.isna(related_vars):
                    continue
                
                # Parse variabel dengan multiple separator dan normalisasi
                lines = parse_related_variables(related_vars)
                
                if not lines:
                    continue
                
                # Filter hanya variabel yang ada di self.components dan max_values
                valid_vars = [var for var in lines if var in components and var in max_values]
                
                # Jika tidak ada variabel yang valid, skip FMEA item ini
                if not valid_vars:
                    continue
                
                # ===== LOGIKA AND YANG BENAR =====
                # SEMUA variabel yang valid harus >= alarm threshold
                all_alarm_or_trip = all(
                    is_alarm_or_trip(max_values[var], components[var]) 
                    for var in valid_vars
                )
                
                if all_alarm_or_trip:
                    excel_row = row_idx + 10
                    display_row = excel_row
                    
                    # Ambil RPN untuk sorting
                    rpn_val = None
                    if "RPN" in df_fmea.columns:
                        try:
                            rpn_val = df_fmea.iloc[row_idx]["RPN"]
                            rpn_num = float(rpn_val) if pd.notna(rpn_val) else float("-inf")
                        except:
                            rpn_num = float("-inf")
                    else:
                        rpn_num = float("-inf")
                    
                    # Kolom yang ingin ditampilkan
                    columns_to_extract = [
                        "LEVEL", "Item Identification", "Failure Mode",
                        "Related Variable(s)", "SEV", "RPN", "Recommended Action",
                        "CAUSE OF CATEGORY", "RECOMMENDATION", "PROCEDURE"
                    ]
                    
                    fmea_item = {
                        "Excel_Row": display_row,
                        "RPN": rpn_val if pd.notna(rpn_val) else "No RPN",
                        "RPN_Numeric": rpn_num
                    }
                    
                    for col in columns_to_extract:
                        if col in df_fmea.columns:
                            value = df_fmea.iloc[row_idx][col]
                            fmea_item[col] = value if pd.notna(value) else "No Information"
                        else:
                            fmea_item[col] = "No Information"
                    
                    triggered_fmea.append(fmea_item)
        
        # Sort triggered FMEA by RPN (descending)
        triggered_fmea.sort(key=lambda x: x["RPN_Numeric"], reverse=True)
        
        # Buat DataFrame untuk triggered FMEA
        df_triggered = pd.DataFrame(triggered_fmea)
        if not df_triggered.empty:
            # Remove RPN_Numeric column (hanya untuk sorting)
            df_triggered = df_triggered.drop(columns=["RPN_Numeric"])
        
        # Determine overall status
        if trip_vars:
            overall_status = "TRIP"
        elif alarm_vars:
            overall_status = "ALARM"
        else:
            overall_status = "NORMAL"
        
        # Summary
        summary = {
            "total_variables": len(max_values),
            "trip_count": len(trip_vars),
            "alarm_count": len(alarm_vars),
            "normal_count": len(max_values) - len(trip_vars) - len(alarm_vars),
            "triggered_fmea_count": len(triggered_fmea),
            "trip_variables": trip_vars,
            "alarm_variables": alarm_vars
        }
        
        return {
            "status_variables": df_status,
            "triggered_fmea": df_triggered,
            "overall_status": overall_status,
            "summary": summary
        }

def evaluate_crisp(value, params):
        """
        Evaluasi crisp logic: return status berdasarkan threshold
        
        Returns:
        --------
        str : "Normal", "Alarm", atau "Trip"
        """
        if value >= params["trip"]:
            return "Trip"
        elif value >= params["alarm"]:
            return "Alarm"
        else:
            return "Normal"
    
def is_alarm_or_trip(value, params):
    """
    Helper function untuk cek apakah nilai >= alarm threshold
    
    Returns:
    --------
    bool : True jika >= alarm, False jika normal
    """
    return value >= params["alarm"]

def save_to_excel(result, output_dir, filename=None):
        """
        Simpan hasil analisis ke file Excel dengan multiple sheets
        
        Parameters:
        -----------
        result : dict
            Output dari process_input()
        filename : str
            Nama file output (opsional, default: "monitoring_report_TIMESTAMP.xlsx")
        
        Returns:
        --------
        str : Path ke file yang disimpan
        """
        # Generate filename jika tidak diberikan
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"monitoring_report_{timestamp}.xlsx"
        
        output_path = output_dir + "/" + filename
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # ===== SHEET 1: STATUS VARIABLES =====
            summary = result["summary"]
            df_status = result["status_variables"]
            
            # Buat summary section
            summary_data = {
                "Metric": [
                    "Overall Status",
                    "Total Variables",
                    "Trip Count",
                    "Alarm Count",
                    "Normal Count",
                    "Triggered FMEA Items"
                ],
                "Value": [
                    result["overall_status"],
                    summary["total_variables"],
                    summary["trip_count"],
                    summary["alarm_count"],
                    summary["normal_count"],
                    summary["triggered_fmea_count"]
                ]
            }
            df_summary_section = pd.DataFrame(summary_data)
            df_summary_section.to_excel(writer, sheet_name="Status Variables", index=False, startrow=0)
            
            # Tulis status variables table dibawahnya
            df_status.to_excel(writer, sheet_name="Status Variables", index=False, startrow=len(df_summary_section) + 2)
            
            # Format sheet Status Variables
            ws_status = writer.sheets["Status Variables"]
            ws_status.column_dimensions['A'].width = 35
            ws_status.column_dimensions['B'].width = 15
            ws_status.column_dimensions['C'].width = 12
            ws_status.column_dimensions['D'].width = 18
            ws_status.column_dimensions['E'].width = 15
            ws_status.column_dimensions['F'].width = 12
            
            # Add color formatting untuk status
            fill_trip = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            fill_alarm = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            fill_normal = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)
            
            status_col = 6 + 1  # Column F (Status)
            for row in range(len(df_summary_section) + 3, len(df_summary_section) + 3 + len(df_status)):
                cell = ws_status.cell(row=row, column=status_col)
                if cell.value == "Trip":
                    cell.fill = fill_trip
                    cell.font = font_white
                elif cell.value == "Alarm":
                    cell.fill = fill_alarm
                    cell.font = font_white
                elif cell.value == "Normal":
                    cell.fill = fill_normal
                    cell.font = font_white
            
            # ===== SHEET 2: TRIGGERED FMEA =====
            df_triggered = result["triggered_fmea"]

                # Add header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            if not df_triggered.empty:
                # save_to_prescriptions(df_triggered)
                df_triggered.to_excel(writer, sheet_name="Triggered FMEA", index=False)
                
                # Format sheet Triggered FMEA
                ws_fmea = writer.sheets["Triggered FMEA"]
                ws_fmea.column_dimensions['A'].width = 12
                ws_fmea.column_dimensions['B'].width = 8
                
                # Set column widths untuk semua kolom
                for col_idx, col_name in enumerate(df_triggered.columns, 1):
                    ws_fmea.column_dimensions[chr(64 + col_idx)].width = 20
                
                for cell in ws_fmea[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Buat sheet kosong dengan pesan
                df_no_trigger = pd.DataFrame({"Message": ["No triggered FMEA items"]})
                df_no_trigger.to_excel(writer, sheet_name="Triggered FMEA", index=False)
            
            # ===== SHEET 3: SUMMARY =====
            summary_data_detail = {
                "Overall Status": [result["overall_status"]],
                "Total Variables": [summary["total_variables"]],
                "Trip Count": [summary["trip_count"]],
                "Alarm Count": [summary["alarm_count"]],
                "Normal Count": [summary["normal_count"]],
                "Triggered FMEA Items": [summary["triggered_fmea_count"]]
            }
            
            # Trip variables list
            if summary["trip_variables"]:
                trip_vars_str = ", ".join(summary["trip_variables"])
            else:
                trip_vars_str = "None"
            
            # Alarm variables list
            if summary["alarm_variables"]:
                alarm_vars_str = ", ".join(summary["alarm_variables"])
            else:
                alarm_vars_str = "None"
            
            summary_data_detail["Trip Variables"] = [trip_vars_str]
            summary_data_detail["Alarm Variables"] = [alarm_vars_str]
            
            df_summary = pd.DataFrame(summary_data_detail)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)
            
            # Format sheet Summary
            ws_summary = writer.sheets["Summary"]
            for col_idx in range(1, len(df_summary.columns) + 1):
                ws_summary.column_dimensions[chr(64 + col_idx)].width = 20
            
            # Add header formatting
            for cell in ws_summary[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        print(f"\n✅ Excel report saved: {output_path}")
        return str(output_path)

def save_to_prescriptions(df):
    records = df.to_dict(orient="records")
    version = datetime.now().strftime("%Y%m%d%H%M")

    for record in records:
        sql = """
            MERGE INTO ugm25_prescriptions t
            USING (
                SELECT :version AS version,
                    :excel_row AS excel_row,
                    :rpn AS rpn,
                    :data_level AS data_level,
                    :item_identification AS item_identification,
                    :failure_mode AS failure_mode,
                    :related_variable AS related_variable,
                    :recommended_action AS recommended_action,
                    :cause_category AS cause_category,
                    :recommendation AS recommendation,
                    :procedure AS procedure
                FROM dual
            ) s
            ON (t.version = s.version AND t.excel_row = s.excel_row)

            WHEN MATCHED THEN
                UPDATE SET
                    t.rpn = s.rpn,
                    t.data_level = s.data_level,
                    t.item_identification = s.item_identification,
                    t.failure_mode = s.failure_mode,
                    t.related_variable = s.related_variable,
                    t.recommended_action = s.recommended_action,
                    t.cause_category = s.cause_category,
                    t.recommendation = s.recommendation,
                    t.procedure = s.procedure,
                    t.updated_at = SYSDATE

            WHEN NOT MATCHED THEN
                INSERT (
                    version, excel_row, rpn, data_level, item_identification,
                    failure_mode, related_variable, recommended_action,
                    cause_category, recommendation, procedure
                )
                VALUES (
                    s.version, s.excel_row, s.rpn, s.data_level, s.item_identification,
                    s.failure_mode, s.related_variable, s.recommended_action,
                    s.cause_category, s.recommendation, s.procedure
                )
            """


        insert = {
            "version": version,
            "excel_row": record["Excel_Row"],
            "rpn": record["RPN"],
            "data_level": record["LEVEL"],
            "item_identification": record["Item identification"],
            "failure_mode": record["Failure mode"],
            "related_variable": record["Related Variable"],
            "recommended_action": record["Recommended Action"],
            "cause_category": record["CAUSE OF CATEGORY"],
            "recommendation": record["RECOMMENDATION"],
            "procedure": record["PROCEDURE"]
        }

        execute_query(sql, insert)

# SELECT
#     sensor_id,
#     MAX(your_value_column) AS max_value
# FROM your_table
# WHERE sensor_id IN (2, 3, 4)
# GROUP BY sensor_id;
def prepare_data(config, task):
    now = datetime.strptime(task['START_AT'].strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S")

    date_to = now.strftime("%Y-%m-%d")
    date_from = (now - timedelta(days=config['TIMESTAMP_DAYS'])).strftime("%Y-%m-%d")
    print('date from ', date_from, ' date to ', date_to)

    sql = "SELECT p.sensor_id, s.name, MAX(p.value) AS max_value FROM " + settings.TABLE_PREDICTIONS + " p LEFT JOIN "+ settings.TABLE_SENSORS +" s on p.SENSOR_ID = s.ID"
    sql += " WHERE RECORD_TIME >= TO_DATE(:date_from, 'YYYY-MM-DD') AND RECORD_TIME < TO_DATE(:date_to, 'YYYY-MM-DD') + INTERVAL '1' DAY"
    sql += " GROUP BY p.sensor_id, s.name"
    params = {"date_from": date_from, "date_to": date_to}
    df = fetch_all(sql, params)

    print(f"data_records_count: {len(df)}")
    return df